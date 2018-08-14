#!/usr/local/bin/python3


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment, PatternFill, Color # Added
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
# from sys import argv
from openpyxl.chart import (
    PieChart3D,
    Reference
)
import pandas as pd
import numpy as np
import glob
import pdb
import os
import getpass
import shutil
import re


district_name = input("What is the name of the District? ")
account_name = getpass.getuser()
make_folder1 = os.mkdir('/Users/' + account_name + '/Desktop/' + district_name)
make_folder2 = os.mkdir('/Users/' + account_name + '/Desktop/' + district_name + '/XLSX/')
files = glob.iglob('/Users/' + account_name + '/Downloads/ScanResult_*.xlsx', recursive=True)
for file in files:
    shutil.copy(file, '/Users/' + account_name + '/Desktop/' + district_name + '/XLSX/')

directory = glob.iglob('/Users/' + account_name + '/Desktop/' + district_name + '/XLSX/*.xlsx', recursive=True)

print("Adding file location to Sheets")
def add_location():
    for f in directory:
        read_file = pd.read_excel(f, header=None)
        site_name = read_file.iloc[4,1]
        if re.search('[0-9][0-9]\s-\sSCHEDULED\s-\s', site_name):
            site_name = re.sub('[0-9][0-9]\s -\sSCHEDULED\s-\s', '', site_name)
        else:
            site_name = site_name.replace('SCHEDULED - ', '')

        read_file.insert(0,'Location',site_name)
        read_file.iloc[[5], [0]] = 'Location'
        #print(site_name)
        read_file.to_excel(f)

add_location()

directory = glob.iglob('/Users/' + account_name + '/Desktop/' + district_name + '/XLSX/*.xlsx', recursive=True)
for f in directory:
    wb = load_workbook(f)
    wb.save(filename=f)
print("Save-As Complete")

print("Grabbing Files")

directory = glob.iglob('/Users/' + account_name + '/Desktop/' + district_name + '/XLSX/*.xlsx', recursive=True)

def merge_files():
    merge_file = pd.DataFrame()
    for f in directory:
    #pdb.set_trace()
        read_file = pd.read_excel(f, header=6) # damn, either 5 or 6, now its all broken
        merge_file = merge_file.append(read_file)
    merge_file.to_excel('/Users/' + account_name + '/Desktop/MergedFile.xlsx')

merge_files()

print("Merging complete")

openfile = pd.read_excel('/Users/' + account_name  + '/Desktop/MergedFile.xlsx')
# - python debugger - pdb.set_trace()
vuln_counts = openfile[(openfile['Risk Level'] == 'Serious') | \
    (openfile['Risk Level'] == 'High') | \
    (openfile['Risk Level'] == 'Medium')]\
    .groupby(['Vulnerability','Risk Level']).size().reset_index(name='Quantity')

vuln_counts.to_excel('/Users/' + account_name + '/Desktop/' + district_name + ' Vuln Count.xlsx')


min_col, min_row, max_col, max_row = range_boundaries("A:P")
wb = load_workbook('/Users/' + account_name + '/Desktop/MergedFile.xlsx')
print("File Loaded")
ws = wb["Sheet1"]

serverity_column = 9 # == J
first_column = 0 # == A

newfile = Workbook()

# This creates new sheets
#def create_sheets():
ws0 = newfile.create_sheet("Executive Summary", 0)
ws1 = newfile.create_sheet("Critical", 1)
ws2 = newfile.create_sheet("High", 2)
ws3 = newfile.create_sheet("Medium", 3)
ws4 = newfile.create_sheet("Low", 4)
ws5 = newfile.create_sheet("Info", 5)

#create_sheets()
#pdb.set_trace()
# Set tab Color
ws1.sheet_properties.tabColor='FF00FF' # Critical
ws2.sheet_properties.tabColor='FF0000' # High
ws3.sheet_properties.tabColor='FFC000' # Medium
ws4.sheet_properties.tabColor='FFFF00' # Low
ws5.sheet_properties.tabColor='00B050' # Info

# Creating Header Row on new sheets, doesn't need to be done on ws5 as that has headers already - don't ask!
def add_header(ws):
    ws['C1'] = "Location"
    ws['D1'] = "Hostname"
    ws['E1'] = "Host IP"
    ws['F1'] = "Service"
    ws['G1'] = "Vuln ID"
    ws['H1'] = "CVSS"
    ws['I1'] = "CVE"
    ws['J1'] = "Risk Level"
    ws['K1'] = "Vulnerability"
    ws['L1'] = "Observation"
    ws['M1'] = "Remediation"
    ws['N1'] = "Consequences"
    ws['O1'] = "Test Output"
    ws['P1'] = "Operating Sytem/Software"
    return ws

def handlews(*sheets):
    return [add_header(sheet) for sheet in sheets]

listofsheets = handlews(*[ws1, ws2, ws3, ws4])


"""
This gets the sheet names of the new file, finds the default created sheet
and removes it.
"""
sheet_names = newfile.sheetnames
find_sheet = newfile['Sheet']
rm_sheet = newfile.remove(find_sheet)

"""
This will iterate through the vulnerability report sheet and copy the cell value
and entire row over to the new sheets.
"""

for row in ws.iter_rows():
    if row[serverity_column].value == 'Serious':
        ws1.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[serverity_column].value == 'High':
        ws2.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[serverity_column].value == 'Medium':
        ws3.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[serverity_column].value == 'Low':
        ws4.append((cell.value for cell in row[min_col-1:max_col]))
    else:
        ws5.append((cell.value for cell in row[min_col-1:max_col]))
"""
This makes the cells have text wrap
"""
for row in ws1.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws3.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws4.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws5.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

# ws1 column dimensions
def format_column(ws):
    ws.column_dimensions["C"].width = 14.50
    ws.column_dimensions["D"].width = 17
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 38
    ws.column_dimensions["M"].width = 25
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["O"].width = 30
    ws.column_dimensions["P"].width = 30
    return ws

def handle_col(*sheets):
    return [format_column(sheet) for sheet in sheets]

proper_width = handle_col(*[ws1, ws2, ws3, ws4, ws5])

print("Processing stuff in the background")
row_count_serious = ws1.max_row - 1
row_count_high = ws2.max_row - 1
row_count_medium = ws3.max_row - 1
row_count_low = ws4.max_row - 1
row_count_info = ws5.max_row - 1

a = ws1['A2']
b = ws2['A2']
c = ws3['A2']
d = ws4['A2']
e = ws5['A2']
ws1.freeze_panes = a
ws2.freeze_panes = b
ws3.freeze_panes = c
ws4.freeze_panes = d
ws5.freeze_panes = e

def hide_column(ws, column_id):
    if isinstance(column_id, int):
        assert column_id >= 1, "Column numbers must be 1 or greater"
        column_id = openpyxl.utils.get_column_letter(column_id)
    column_dimension = ws.column_dimensions[column_id]
    column_dimension.hidden = True

def spy():
    hide_column(ws1, 1)
    hide_column(ws2, 1)
    hide_column(ws3, 1)
    hide_column(ws4, 1)
    hide_column(ws5, 1)
    hide_column(ws1, 2)
    hide_column(ws2, 2)
    hide_column(ws3, 2)
    hide_column(ws4, 2)
    hide_column(ws5, 2)
spy()

# SANITY CHECK
#wb2 = newfile.get_sheet_names()
#print(wb2)
wb2 =newfile.active

print("Creating Pie Chart")
data = [
    ['Severity', 'Quantity'],
    ['Critical', row_count_serious],
    ['High', row_count_high],
    ['Medium', row_count_medium],
    ['Low', row_count_low],
    #['Info', row_count_info]
]
for row in data:
        ws0.append(row)

pie = PieChart3D()
labels = Reference(ws0, min_col=1, min_row=2, max_row=5)
data = Reference(ws0, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Vulnerabilities by Severity"

ws0.add_chart(pie, "D1")

outpath = newfile.save('/Users/' + account_name + '/Desktop/' + district_name + ' Final Report.xlsx')
print("Your files are saved on your Desktop, it is called " + district_name + ' Final Report and ' + district_name
      + " Vuln Count")

os.remove('/Users/' + account_name + '/Desktop/MergedFile.xlsx')
shutil.rmtree('/Users/' + account_name + '/Desktop/' + district_name)