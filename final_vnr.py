#!/usr/bin/env python3

import mysql.connector
import pandas as pd
from sys import exit
from subprocess import Popen
import openpyxl
from openpyxl import load_workbook
from os import remove
from getpass import getuser
from datetime import datetime
# import pdb
import final_vnr_pub_date as fvpd

"""
Ths script is setup to run from a Mac and Mac only!
"""

start_time = datetime.now()
# This is a dependency, we are creating a background ssh tunnel, this is to be able to access the AlienVault DB
# The REQUIREMENT is the your public key is in the known_hosts on sec-mobile-one, use ssh-copy-id root@sec-mobile-one
# and that you also have the proper settings in your ssh config
print("Starting background process. ")
ssh_magic = Popen(['ssh', '-N', '-L', '3306:127.0.0.1:3306', 'sec-mobile-one'], shell=False)

if ssh_magic.returncode == 1:
    print("Killing background process cause you suck at coding!")
    ssh_magic.kill()

# This is to save the file to your profile and  for the name of the file that is created.
current_user = getuser()
district_name = input("What is the name of the district? ")

# This is the query to get the Vulnerability Results from the Database
query = ("""SELECT vnreps.name as Location, hostname as Hostname, hostip as 'Host IP', service as Service, 
        scriptid as 'Vuln ID', vnp.cve_id as CVEs, vnres.risk as 'Risk Level', vnp.name, vnf.name, vnc.name,
        vnp.copyright, vnp.version, msg 
        FROM alienvault.vuln_nessus_results vnres 
        join vuln_nessus_reports vnreps on vnreps.report_id = vnres.report_id
        join vuln_nessus_plugins vnp on vnres.scriptid = vnp.id
        join vuln_nessus_category vnc on vnc.id = vnp.category
        join vuln_nessus_family vnf on vnf.id = vnp.family
        order by vnres.risk ASC""")

# This is to get the total numbers per site as scans are based on physical locations
query1 = ("""SELECT REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(vjobs.name, "SCHEDULED - ", "")
            , " 1", "") , " 2", "") , " 3", "")  , " 4", "")  , " 5", "")  , " 6", "")  , " 7", "")  , " 8", "")  , " 9", "")  as Location, 
            SUM(vSerious) as Critical, SUM(vHigh) as High, SUM(vMed) as Medium, SUM(vLow) as Low from vuln_nessus_report_stats vnrstats
            join vuln_jobs vjobs on vjobs.id = vnrstats.name
            group by Location""")

# This is to get the device count
query2 = ("SELECT distinct(count(hostname)) FROM alienvault.host")

# Accessing the AlienVault MySQL Database
def db_call(query):
    try:
        cnx = mysql.connector.connect(user='svcpython', password='e6$E7sQBUegYayoU',
                                      host='localhost',
                                      database='alienvault')
    except Exception as e:
        print("I am unable to connect to the MySQL Instance: " + str(e))
        exit(1)


    cursor = cnx.cursor()
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    cnx.close()
    print("Retrieving results from the Database.")
    return results


# GLOBAL
header_row = ['Location', 'Hostname', 'Host IP', 'Service', 'Vuln ID', 'CVE', 'Risk Level', 'Vulnerability',
              '9', '10', '11', '12', 'Blob']
header_row1 = ['Location', 'Critical', 'High', 'Medium', 'Low']
df = pd.DataFrame(db_call(query), columns=header_row)
df = df.astype(str)
df1 = pd.DataFrame(db_call(query1), columns=header_row1)
df_asset_count = pd.DataFrame(db_call(query2))
asset_count = df_asset_count[0][0]
df2 = pd.DataFrame(df[['Vulnerability', 'Risk Level']])
df4 = pd.DataFrame(df[['Host IP', 'Vulnerability', 'Risk Level']])
df4['Risk Level'].replace(to_replace=['1', '2', '3', '4', '5', '6', '7', '8'],
                              value=['Critical', 'High', 'Medium', 'Medium/Low', 'Low/Medium', 'Low', 'Info',
                                     'Exceptions'], inplace=True)
df4 = df4[df4['Risk Level'] != 'Info']
df4 = df4[df4['Risk Level'] != 'Low']
host_vuln_size = df4.groupby(['Host IP', 'Risk Level']).size().reset_index(name='Quantity')
host_vuln_size.sort_values(by=['Risk Level','Quantity'], ascending=[True, False], inplace=True)

top10 = host_vuln_size.head(n=10)
top10_parsed = top10[['Host IP', 'Quantity']]
tlist = top10_parsed.values.tolist()
"""
# FIX THIS
top5vulns = df4.head(n=5)
t5vlist = top5vulns.values.tolist()


rce1 = df3[df3['Vulnerability'].str.contains('Remote Code Execution')]
rce2 = df3[df3['Vulnerability'].str.contains('RCE')]
rce = rce1.append(rce2)

# pdb.set_trace()

rce.sort_values(by=['Risk Level', 'Quantity'], ascending=[True], inplace=True)

df5 = pd.DataFrame(rce)
rcetop10 = pd.DataFrame(rce.head(n=10))
rcetop10['Risk Level'].replace('Critical', 10, inplace=True)

rcetop10lst = rcetop10.values.tolist()

df5['Risk Level'].replace(10, 'Critical', inplace=True)
# df5.drop('Quantity', axis=1, inplace=True)
"""
def df_rm_garbage():
    # The following function removes beginning digits (1 - 10+), 'SCHEDULED - ' and any trailing digits (1 - 10+)
    # in the Locations Column of the DataFrame
    print("Removing 'SCHEDULED - ' and number from the Location")
    df['Location'] = df['Location'].replace(to_replace=r'SCHEDULED - ', value='', regex=True)
    df['Location'] = df['Location'].replace(to_replace=r'(\d+)', value='', regex=True)
    df['Location'] = df['Location'].replace(to_replace=r'( - )', value='', regex=True)
    df1['Location'] = df1['Location'].replace(to_replace=r'SCHEDULED - ', value='', regex=True)
    df1['Location'] = df1['Location'].replace(to_replace=r'(\d+)', value='', regex=True)
    # df1['Location'] = df1['Location'].replace(to_replace=r'( - )', value='', regex=True)
    return df, df1


def df_vuln_counts():
    # The following replaces the the numerical structure of the 'Risk Level' column with a string based structure.
    # It also groups vulnerabilites by quantity in order to determine what the most prevalent vulns are within the
    # environment. It finally sorts it based on alphabetical order first then by quantity in descending fashion.
    df2['Risk Level'].replace(to_replace=['1', '2', '3', '4', '5', '6', '7', '8'],
                              value=['Critical', 'High', 'Medium', 'Medium/Low', 'Low/Medium', 'Low', 'Info',
                                     'Exceptions'], inplace=True)
    df3 = df2[(df2['Risk Level'] == 'Critical') | (df2['Risk Level'] == 'High') | \
              (df2['Risk Level'] == 'Medium')].groupby(['Vulnerability', 'Risk Level']) \
        .size().reset_index(name='Quantity')
    df3.sort_values(by=['Risk Level', 'Quantity'], ascending=[True, False], inplace=True)
    # df3.drop(df3[df3.Quantity <= 10].index, inplace=True)
    return df3


def df_shenanigans():
    print("Replacing 'Risk Level' integer with string.")
    print("Creating AlienVault Excel view for Columns.")
    df['Vulnerability'] = df['Vulnerability'] + '\n' + 'Family name: ' + df['9'] + '\n' + 'Category: ' + df['10'] + \
                          '\n' + 'Copyright: ' + df['11'] + '\n' + 'Version: ' + df['12']
    df['Risk Level'].replace(to_replace=['1', '2', '3', '4', '5', '6', '7', '8'],
                             value=['Critical', 'High', 'Medium', 'Medium/Low', 'Low/Medium', 'Low', 'Info',
                                    'Exceptions'], inplace=True)
    return df


def df_extract():
    # Extracting values from the AlienVault Database as it is a hodgepodge of information, this presents the information
    # like when you download the Excel
    print("Extracting fields and placing the data in AlienVault Excel view for Columns")
    df['CVSS'] = df['Blob'].str.extract('Score:\s(\d.+)', expand=True)
    df['Observation'] = df['Blob'].str.extract('Summary:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['Remediation'] = df['Blob'].str.extract('Solution:\n\n((?:[^\n][\n]?)+)', expand=False)
    insight_df = df['Blob'].str.extract('(Insight:\n\n(.|\n)*)', expand=False)
    df_insight = insight_df.drop([1], axis=1)
    df_insight = df_insight.replace('(Vulnerability Detection Method:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(CVSS Base Vector:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(Affected Software\/OS:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(CVSS Base Score:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(Solution:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(References:(.|\n)*)', '', regex=True)  # YES
    df_insight = df_insight.replace('(Summary:(.|\n)*)', '', regex=True) # YES
    df_insight = df_insight.replace('(Impact:(.|\n)*)', '', regex=True)
    df_insight = df_insight.replace('(&#039;)', '', regex=True)  # YES
    df_insight = df_insight.replace('\n', ' ', regex=True)
    df_insight = df_insight.replace(' - ', '\n - ', regex=True)
    df['insight']= df_insight
    df_impact = df['Blob'].str.extract('(Impact:\n\n((?:[^\n][\n]?)(.|\n)*))', expand=False)
    new_df = df_impact.drop([1, 2], axis=1)
    new_df = new_df.replace('(Vulnerability Detection Method:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(CVSS Base Vector:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(Affected Software\/OS:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(Solution:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(References:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(Insight:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(Summary:(.|\n)*)', '', regex=True)
    new_df = new_df.replace('(&#039;)', "'", regex=True)
    new_df = new_df.replace('\n', ' ', regex=True)
    df['impact'] = new_df
    df['references'] = df['Blob'].str.extract('(References:\n\n(?:[^\n][\n]?)+)', expand=False)
    df['references'] = df['references'].replace('\n', ' ', regex=True)
    df['references'] = df['references'].replace('References:', 'References:\n', regex=True)
    df['Consequences'] = df['insight'] + '\n\n' + df['impact'] + '\n\n' + df['references']
    df['Test Output'] = df['Blob'].str.extract('Result:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['Operating System/Software'] = df['Blob'].str.extract('OS:\n\n((?:[^\n][\n]?)+)', expand=False)
    return df


def df_drop_columns():
    # Removing Columns that are no longer necessary
    print("Cleaning up the table structure. ")
    df.drop(['9', '10', '11', '12'], inplace=True, axis=1)
    df.drop(['Blob', 'insight', 'references'], inplace=True, axis=1)
    return df


def df_apply_sort():
    # This sort the DataFrame by CVSS Score. 10 - 0 in descending order. This is for precedence sake
    print("Sorting results by CVSS.")
    df[['CVSS']] = df[['CVSS']].apply(pd.to_numeric)
    df.sort_values(by=['CVSS', 'Host IP'], ascending=[False, False], inplace=True)
    return df


def output_df():
    ordered_df = df[
        ['Location', 'Hostname', 'Host IP', 'Service', 'Risk Level', 'CVSS', 'Vuln ID', 'CVE', 'Vulnerability',
         'Observation', 'Remediation', 'Consequences', 'Test Output', 'Operating System/Software']]
    return ordered_df


def df_sheets():
    s = output_df()
    critical_df = s.where(df['Risk Level'] == 'Critical')
    high_df = s.where(df['Risk Level'] == 'High')
    medium_df = s.where(df['Risk Level'] == 'Medium')
    low_df = s.where(df['Risk Level'] == 'Low')
    info_df = s.where(df['Risk Level'] == 'Info')
    return critical_df, high_df, medium_df, low_df, info_df


def rm_nans():
    s = df_sheets()
    processed_crits = s[0].dropna(how='all')
    processed_highs = s[1].dropna(how='all')
    processed_mediums = s[2].dropna(how='all')
    processed_lows = s[3].dropna(how='all')
    processed_infos = s[4].dropna(how='all')
    return processed_crits, processed_highs, processed_mediums, processed_lows, processed_infos


def df_v_counts():
    s = rm_nans()
    crits = s[0].shape[0]
    highs = s[1].shape[0]
    meds = s[2].shape[0]
    lows = s[3].shape[0]
    return crits, highs, meds, lows


def ex_data():
    from xlsxwriter import Workbook
    s = df_v_counts()
    t = df_vuln_counts()
    print("Adding Charts to the Executive Summary Sheet")
    file = Workbook('/Users/' + current_user + '/Desktop/Initial Report.xlsx')
    ws0 = file.add_worksheet('Executive Summary')
    ws8 = file.add_worksheet('Chart Data')

# ====== CUMULATIVE CHART ========== #
    headings = ['Severity', 'Quantity']
    data = [
        ['Critical', 'High', 'Medium', 'Low'],
        [s[0], s[1], s[2], s[3]]
    ]
    ws8.write_row('A1', headings)
    ws8.write_column('A2', data[0])
    ws8.write_column('B2', data[1])

    chart1 = file.add_chart({'type': 'bar'})

    # Configure the series and add user defined segment colors.
    chart1.add_series({
        # 'name': 'Vulnerabiliities by Severity',
        # CHANGED VALUES, SHOULD WORK NOW
        'categories': "='Chart Data'!$A$2:$A$5",
        'values': "='Chart Data'!$B$2:$B$5",
        'data_labels': {'value': True},
        'points': [
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF0000'}},
            {'fill': {'color': '#FFC000'}},
            {'fill': {'color': '#FFFF00'}}
        ],
    })

    # Add a title.
    chart1.set_title({'name': 'Vulnerabilities by Severity'})
    chart1.set_x_axis({'name': 'Quantity'})
    chart1.set_y_axis({'name': 'Severity'})

    # Insert the chart into the worksheet (with an offset).
    ws0.insert_chart('A1', chart1, {'x_offset': 25, 'y_offset': 10})

# ==================================================================== #
# ======= CHART PER LOCATION ========== #
    Quantity = df1.shape[0]
    df_list = df1.values.tolist()

    Location = [item[0] for item in df_list]
    Critical = [item[1] for item in df_list]
    High = [item[2] for item in df_list]
    Medium = [item[3] for item in df_list]
    Low = [item[4] for item in df_list]

    # Converting from strings to integers
    Critical = list(map(int, Critical))
    High = list(map(int, High))
    Medium = list(map(int, Medium))
    Low = list(map(int, Low))

    headings = ['Facility', 'Critical', 'High', 'Medium', 'Low']
    data = [
        Location,
        Critical,
        High,
        Medium,
        Low
    ]

    ws8.write_row('D1', headings)
    ws8.write_column('D2', data[0])
    ws8.write_column('E2', data[1])
    ws8.write_column('F2', data[2])
    ws8.write_column('G2', data[3])
    ws8.write_column('H2', data[4])

    for i in range(Quantity):
        chart2 = file.add_chart({'type': 'bar'})

        chart2.add_series({
            'name': Location[i],
            # THIS SHOULD WORK
            'categories': ['Chart Data', 0, 4, 0, 7],
            'values': ['Chart Data', i + 1, 4, i + 1, 7],
            'data_labels': {'value': True},
            'points': [
                {'fill': {'color': '#FF00FF'}},
                {'fill': {'color': '#FF0000'}},
                {'fill': {'color': '#FFC000'}},
                {'fill': {'color': '#FFFF00'}}
            ],
        })
        # chart.set_title({'name': 'Location by Severity'})
        chart2.set_x_axis({'name': 'Quantity'})
        chart2.set_y_axis({'name': 'Severity'})

        chart2.set_style(11)

        ws0.insert_chart('K' + str(17 * i + 2), chart2, {'x_offset': 25, 'y_offset': 10 * (i + 1)})

# ================================================================================= #
# ==============TOP 10 HOSTS CHART============= #
    host_ip = [item[0] for item in tlist]
    quantity = [item[1] for item in tlist]
    headings = ['Host IP', 'Quantity']
    data = [
        host_ip,
        quantity
    ]

    ws8.write_row('L1', headings)
    ws8.write_column('L2', data[0])
    ws8.write_column('M2', data[1])

    chart3 = file.add_chart({'type': 'column'})

    # Configure the series and add user defined segment colors.
    chart3.add_series({
        # 'name': 'Vulnerabilities by Severity',
        # CHANGED VALUES, SHOULD WORK NOW
        'categories': "='Chart Data'!$L$2:$L$11",
        'values': "='Chart Data'!$M$2:$M$11",
        'data_labels': {'value': True},
        'points': [
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
        ],

    })

    # Add a title.
    chart3.set_legend({'position': 'none'})
    chart3.set_title({'name': 'Critical: Top 10 Hosts'})
    chart3.set_x_axis({'name': 'Host IP'})
    chart3.set_y_axis({'name': 'Quantity'})

    ws0.insert_chart('A17', chart3, {'x_offset': 25, 'y_offset': 10})

# ========================================================================== #
# ============= TOP 5 COMMON VULNERABILITIES CHART ============= #
    top5vulns = t.head(n=5)
    t5vlist = top5vulns.values.tolist()

    vuln = [item[0] for item in t5vlist]
    quantity = [item[2] for item in t5vlist]
    headings = ['Vulnerability', 'Quantity']
    data = [
        vuln,
        quantity
    ]

    ws8.write_row('O1', headings)
    ws8.write_column('O2', data[0])
    ws8.write_column('P2', data[1])

    chart4 = file.add_chart({'type': 'bar'})

    # Configure the series and add user defined segment colors.
    chart4.add_series({
        # 'name': 'Vulnerabilities by Severity',
        # CHANGED VALUES, SHOULD WORK NOW
        'categories': "='Chart Data'!$O$2:$O$6",
        'values': "='Chart Data'!$P$2:$P$6",
        'data_labels': {'value': True},
        'points': [
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
        ],
    })

    # Add a title.
    chart4.set_legend({'position': 'none'})
    chart4.set_title({'name': 'Common Enterprise Vulnerabilities'})
    chart4.set_x_axis({'name': 'Quantity'})
    chart4.set_y_axis({'name': 'Vulnerability'})

    ws0.insert_chart('A33', chart4, {'x_offset': 25, 'y_offset': 10})

# ====================================================================== #
# ============= TOP 10 RCE CHART ============= #
    rce1 = t[t['Vulnerability'].str.contains('Remote Code Execution')]
    rce2 = t[t['Vulnerability'].str.contains('RCE')]
    rce = rce1.append(rce2)

    # pdb.set_trace()

    rce.sort_values(by=['Risk Level', 'Quantity'], ascending=[True, False], inplace=True)
    global df5
    df5 = pd.DataFrame(rce)
    rcetop10 = pd.DataFrame(rce.head(n=10))
    rcetop10['Risk Level'].replace('Critical', 10, inplace=True)

    rcetop10lst = rcetop10.values.tolist()

    df5['Risk Level'].replace(10, 'Critical', inplace=True)
    df5.drop('Quantity', axis=1, inplace=True)

    rce = [item[0] for item in rcetop10lst]
    qty = [item[1] for item in rcetop10lst]
    headings = ['Vulnerability', 'Risk Level']
    data5 = [
            rce,
            qty
            ]

    ws8.write_row('R1', headings)
    ws8.write_column('R2', data5[0])
    ws8.write_column('S2', data5[1])

    chart5 = file.add_chart({'type': 'bar'})

# Configure the series and add user defined segment colors.
    chart5.add_series({
        # 'name': 'Vulnerabilities by Severity',
        # CHANGED VALUES, SHOULD WORK NOW
        'categories': "='Chart Data'!$R$2:$R$11",
        'values': "='Chart Data'!$S$2:$S$11",
        'points': [
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF00FF'}},
            ],
        })

    # Add a title.
    chart5.set_size({'width': 1000, 'height': 550})
    chart5.set_legend({'position': 'none'})
    chart5.set_title({'name': 'Risk by Exploitability'})
    chart5.set_x_axis({'name': 'CVSS'})
    chart5.set_y_axis({'name': 'Vulnerability'})
    
    ws0.insert_chart('A88', chart5, {'x_offset': 25, 'y_offset': 10})

    file.close()
    return df5


def writing_to_workbook():
    # This needs to happen after ex_data
    # works = self.rm_nans()
    s = rm_nans()
    print("Writing Data sheets to initial file")
    book = load_workbook('/Users/' + current_user + '/Desktop/Initial Report.xlsx')
    writer = pd.ExcelWriter('/Users/' + current_user + '/Desktop/Initial Report.xlsx', engine='openpyxl', mode='a')
    s[0].to_excel(writer, sheet_name='Critical')
    s[1].to_excel(writer, sheet_name='High')
    s[2].to_excel(writer, sheet_name='Medium')
    s[3].to_excel(writer, sheet_name='Low')
    s[4].to_excel(writer, sheet_name='Info')
    # df.to_excel(writer, sheet_name='Raw Data')
    df1.to_excel(writer, sheet_name='Vulnerabilities by Location')
    df_vuln_counts().to_excel(writer, sheet_name='Unique by Severity')
    host_vuln_size.to_excel(writer, sheet_name='Vulnerability Count by Host')
    ex_data().to_excel(writer, sheet_name='Risk by Exploitability')
    writer.save()
    writer.close()


def final_file():
    final_wb = load_workbook('/Users/' + current_user + '/Desktop/Initial Report.xlsx')
    print("Finalizing file and saving")

    ws_ES = final_wb['Executive Summary']
    ws1 = final_wb['Critical']
    ws2 = final_wb['High']
    ws3 = final_wb['Medium']
    ws4 = final_wb['Low']
    ws5 = final_wb['Info']
    ws6 = final_wb['Vulnerabilities by Location']
    ws7 = final_wb['Unique by Severity']
    ws8 = final_wb['Chart Data']
    ws9 = final_wb['Vulnerability Count by Host']
    ws10 = final_wb['Risk by Exploitability']


    # This moves the Raw Numbers sheet to the end
    sheets = final_wb._sheets
    chart_data = sheets.pop(1)
    sheets.insert(10, chart_data)

    # This adds the device count and quantity to the Raw Numbers sheet
    asset_title = ws8['J1']
    asset_title.value = 'Device Count'
    asset_quantity = ws8['J2']
    asset_quantity.value = asset_count

    # Setting the TAB COLOR
    ws1.sheet_properties.tabColor = 'FF00FF'  # Critical
    ws2.sheet_properties.tabColor = 'FF0000'  # High
    ws3.sheet_properties.tabColor = 'FFC000'  # Medium
    ws4.sheet_properties.tabColor = 'FFFF00'  # Low
    ws5.sheet_properties.tabColor = '00B050'  # Info

    # Iterating over vuln data in order to give the cells textwrap
    vuln_data = [ws1, ws2, ws3, ws4, ws5]

    for ws in vuln_data:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)

    def format_column(ws):
        # This sets the column dimensions for ws1 - ws5
        ws.column_dimensions["B"].width = 14.50
        ws.column_dimensions["C"].width = 17
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 9
        ws.column_dimensions["G"].width = 5
        ws.column_dimensions["H"].width = 8  # was 14
        ws.column_dimensions["I"].width = 14  # was 13
        ws.column_dimensions["J"].width = 25
        ws.column_dimensions["K"].width = 38
        ws.column_dimensions["L"].width = 25
        ws.column_dimensions["M"].width = 25
        ws.column_dimensions["N"].width = 30
        ws.column_dimensions["O"].width = 30
        return ws

    def handle_col(*sheets):
        return [format_column(sheet) for sheet in sheets]

    proper_width = handle_col(*[ws1, ws2, ws3, ws4, ws5])

    ws6.column_dimensions["B"].width = 25
    ws7.column_dimensions["B"].width = 90
    ws8.column_dimensions["D"].width = 15
    ws8.column_dimensions["J"].width = 12
    ws8.column_dimensions["L"].width = 12.50
    ws8.column_dimensions["M"].width = 6.83
    ws8.column_dimensions["O"].width = 65
    ws8.column_dimensions["P"].width = 6.83
    ws8.column_dimensions["R"].width = 65.50
    ws8.column_dimensions["S"].width = 65.50
    ws9.column_dimensions["B"].width = 13
    ws9.column_dimensions["C"].width = 8
    ws10.column_dimensions["B"].width = 100
    ws10.column_dimensions["C"].width = 7.50

    a = ws1['A2']
    b = ws2['A2']
    c = ws3['A2']
    d = ws4['A2']
    e = ws5['A2']
    f = ws6['A2']
    g = ws7['A2']
    h = ws9['A2']
    j = ws10['A2']
    ws1.freeze_panes = a
    ws2.freeze_panes = b
    ws3.freeze_panes = c
    ws4.freeze_panes = d
    ws5.freeze_panes = e
    ws6.freeze_panes = f
    ws7.freeze_panes = g
    ws9.freeze_panes = h
    ws10.freeze_panes = j

    def hide_column(ws, column_id):
        # This is the function creation to hide the dataframe column.
        if isinstance(column_id, int):
            assert column_id >= 1, "Column numbers must be 1 or greater"
            column_id = openpyxl.utils.get_column_letter(column_id)
        column_dimension = ws.column_dimensions[column_id]
        column_dimension.hidden = True

    def spy():
        # This function actually hides the column.
        hide_column(ws1, 1)
        hide_column(ws2, 1)
        hide_column(ws3, 1)
        hide_column(ws4, 1)
        hide_column(ws5, 1)
        hide_column(ws6, 1)
        hide_column(ws7, 1)
        hide_column(ws9, 1)
        hide_column(ws10, 1)

    spy()

    final_wb.save('/Users/' + current_user + '/Desktop/' + district_name + ' Technical Report.xlsx')
    print("Your file is saved to your Desktop. It is called " + district_name + " Technical Report.xslx")
    final_wb.close()


def main():
    df_rm_garbage()
    df_vuln_counts()
    df_shenanigans()
    df_extract()
    df_drop_columns()
    df_apply_sort()
    output_df()
    df_sheets()
    rm_nans()
    fvpd.df_manipulation()
    fvpd.removal_parsing()
    fvpd.more_parsing()
    fvpd.sql_data()
    fvpd.cleaning_dates()
    fvpd.delta_calculation()
    fvpd.counter()
    ex_data()
    writing_to_workbook()
    final_file()


if __name__ == "__main__":
    main()

rm_initial_report = remove('/Users/' + current_user + '/Desktop/Initial Report.xlsx')

ssh_magic.kill()
print("The background process is now terminated!")
end_time = datetime.now()
running_time = (end_time - start_time)
ptime = running_time[2:6]
print("It took " + str(ptime) + " minutes to process.")
