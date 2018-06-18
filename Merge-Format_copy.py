#!/usr/bin/env python3

import glob
import pdb
import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment, PatternFill, Color # Added
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
# from sys import argv
from openpyxl.chart import (
    PieChart3D,
    Reference
)
import pandas as pd
import numpy as np


__author__ = "Issam Beik, Erik Tingelstad"
__version__ = "0.2"

district_name = input("What is the name of the District? ")
# HEADER_ROW = 5
# script = argv
directory = glob.iglob('/Users/beik/Desktop/**/XLSX/*.xlsx', recursive=True)
#directory = glob.iglob('/Users/beik/Desktop/**/XLSX/*.xlsx')
print("Adding file location to Sheets")
def add_location():
    for f in directory:
        read_file = pd.read_excel(f, header=None)
        site_name = read_file.iloc[4,1]
        site_name = re.sub(r'^.*SCHEDULED - ', "", site_name)
        read_file.insert(0,'Location',site_name)
        read_file.iloc[[5], [0]] = 'Location' # This is for the header row in order to merge the documents properly
        #print(site_name)
        os.rename(f, '/Users/beik/Desktop/' + district_name + '/XLSX/' + site_name + '.xlsx') # Figure out a way to put this in its own function
        read_file.to_excel(f)

add_location()
"""
directory = glob.iglob('/Users/beik/Desktop/**/XLSX/*.xlsx', recursive=True)
for f in directory:
    wb = load_workbook(f)
    wb.save(filename=f)
print("Save-As Complete")
"""
print("Grabbing Files")

directory = glob.iglob('/Users/beik/Desktop/**/XLSX/*.xlsx', recursive=True)

def merge_files():
    merge_file = pd.DataFrame()
    for f in directory:
    #pdb.set_trace()
        read_file = pd.read_excel(f, header=6)
        merge_file = merge_file.append(read_file)
    merge_file.to_excel('/Users/beik/Desktop/' + district_name + ' MergedFile.xlsx')

merge_files()

print("Merging complete")

openfile = pd.read_excel('/Users/beik/Desktop' + district_name + ' MergedFile.xlsx')
# - python debugger - pdb.set_trace()
vuln_counts = openfile[(openfile['Risk Level'] == 'Serious') | \
    (openfile['Risk Level'] == 'High') | \
    (openfile['Risk Level'] == 'Medium')]\
    .groupby(['Vulnerability','Risk Level']).size().reset_index(name='counts')

vuln_counts.to_excel('/Users/beik/Desktop/' + district_name + ' Vuln Count.xlsx')
# FIND OUT HOW TO GET THIS TO COPY TO NEWFILE

min_col, min_row, max_col, max_row = range_boundaries("A:P")
wb = load_workbook('/Users/beik/Desktop/' + district_name + ' MergedFile.xlsx')
print("File Loaded")
ws = wb.get_sheet_by_name("Sheet1")

SEVERITY_COLUMN = 9 # == J
CRITICAL_COLOR = 'FF00FF'
HIGH_COLOR = 'FF0000'
MEDIUM_COLOR = 'FFC000'
LOW_COLOR = 'FFFF00'
INFO_COLOR = '00B050'


newfile = Workbook()

# This creates new sheets
#def create_sheets():
ws_ER = newfile.create_sheet("Executive Summary", 0)
ws_critical = newfile.create_sheet("Critical", 1)
ws_high = newfile.create_sheet("High", 2)
ws_medium = newfile.create_sheet("Medium", 3)
ws_low = newfile.create_sheet("Low", 4)
ws_info = newfile.create_sheet("Info", 5)

DATA_WS = [ws_critical, ws_high, ws_medium, ws_low, ws_info]
"""
This gets the sheet names of the new file, finds the default created sheet
and removes it.
"""
sheet_names = newfile.get_sheet_names()
find_sheet = newfile.get_sheet_by_name('Sheet')
rm_sheet = newfile.remove_sheet(find_sheet)

#create_sheets()
#pdb.set_trace()
# Set tab Color
ws_critical.sheet_properties.tabColor=CRITICAL_COLOR
ws_high.sheet_properties.tabColor=HIGH_COLOR
ws_medium.sheet_properties.tabColor=MEDIUM_COLOR
ws_low.sheet_properties.tabColor=LOW_COLOR
ws_info.sheet_properties.tabColor=INFO_COLOR

# Creating Header Row on new sheets, doesn't need to be done on ws_info as that has headers already - don't ask!
def add_header(ws):
    ws['C1'] = "Location"
    ws['D1'] = "Hostname"
    ws['E1'] = "Host IP"
    ws['F1'] = "Service"
    ws['G1'] = "Vuln ID"
    ws['H1'] = "CVSS"
    ws['I1'] = "CVE"
    ws['J1'] = "Risk Level"
    ws['K1'] = "Vulnerability"
    ws['L1'] = "Observation"
    ws['M1'] = "Remediation"
    ws['N1'] = "Consequences"
    ws['O1'] = "Test Output"
    ws['P1'] = "Operating Sytem/Software"
    return ws

def handlews(*sheets):
    return [add_header(sheet) for sheet in sheets]

listofsheets = handlews(*DATA_WS)

"""
This will iterate through the vulnerability report sheet and copy the cell value
and entire row over to the new sheets.
"""

for row in ws.iter_rows():
    if row[SEVERITY_COLUMN].value == 'Serious':
        ws_critical.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[SEVERITY_COLUMN].value == 'High':
        ws_high.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[SEVERITY_COLUMN].value == 'Medium':
        ws_medium.append((cell.value for cell in row[min_col-1:max_col]))
    elif row[SEVERITY_COLUMN].value == 'Low':
        ws_low.append((cell.value for cell in row[min_col-1:max_col]))
    else:
        ws_info.append((cell.value for cell in row[min_col-1:max_col]))
"""
This makes the cells have text wrap
"""
for ws in DATA_WS:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment =  cell.alignment.copy(wrapText=True)
"""
Merge-Format_copy.py:158: DeprecationWarning: Call to deprecated function copy (Use copy(obj) or cell.obj = cell.obj + other).
  cell.alignment =  cell.alignment.copy(wrapText=True)
"""
"""
for row in ws_critical.iter_rows():
    for cell in row:
        cell.alignment =  cell.alignment.copy(wrapText=True) # Tested

for row in ws_high.iter_rows():
    for cell in row:
        cell.alignment =  cell.alignment.copy(wrapText=True) # Tested

for row in ws_medium.iter_rows():
    for cell in row:
        cell.alignment =  cell.alignment.copy(wrapText=True) # Tested

for row in ws_low.iter_rows():
    for cell in row:
        cell.alignment =  cell.alignment.copy(wrapText=True) # Tested

for row in ws_info.iter_rows():
    for cell in row:
        cell.alignment =  cell.alignment.copy(wrapText=True) # Tested
"""
# Formatting Column Dimensions
def format_column(ws):
    ws.column_dimensions["C"].width = 14.50
    ws.column_dimensions["D"].width = 17
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 38
    ws.column_dimensions["M"].width = 25
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["O"].width = 30
    ws.column_dimensions["P"].width = 30
    return ws

def handle_col(*sheets):
    return [format_column(sheet) for sheet in sheets]

proper_width = handle_col(*DATA_WS)

print("Processing stuff in the background")
row_count_serious = ws_critical.max_row - 1
row_count_high = ws_high.max_row - 1
row_count_medium = ws_medium.max_row - 1
row_count_low = ws_low.max_row - 1
row_count_info = ws_info.max_row - 1

a = ws_critical['A2']
b = ws_high['A2']
c = ws_medium['A2']
d = ws_low['A2']
e = ws_info['A2']
ws_critical.freeze_panes = a
ws_high.freeze_panes = b
ws_medium.freeze_panes = c
ws_low.freeze_panes = d
ws_info.freeze_panes = e

def hide_column(ws, column_id):
    if isinstance(column_id, int):
        assert column_id >= 1, "Column numbers must be 1 or greater"
        column_id = openpyxl.utils.get_column_letter(column_id)
    column_dimension = ws.column_dimensions[column_id]
    column_dimension.hidden = True

def spy():
    hide_column(ws_critical, 1)
    hide_column(ws_high, 1)
    hide_column(ws_medium, 1)
    hide_column(ws_low, 1)
    hide_column(ws_info, 1)
    hide_column(ws_critical, 2)
    hide_column(ws_high, 2)
    hide_column(ws_medium, 2)
    hide_column(ws_low, 2)
    hide_column(ws_info, 2)
spy()

# SANITY CHECK
#wb2 = newfile.get_sheet_names()
#print(wb2)
wb2 =newfile.active

print("Creating Pie Chart")
data = [
    ['Severity', 'Quantity'],
    ['Critical', row_count_serious],
    ['High', row_count_high],
    ['Medium', row_count_medium],
    ['Low', row_count_low],
    #['Info', row_count_info]
]
for row in data:
        ws_ER.append(row)

pie = PieChart3D()
labels = Reference(ws_ER, min_col=1, min_row=2, max_row=5)
data = Reference(ws_ER, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Vulnerabilities by Severity"

ws_ER.add_chart(pie, "D1")
outpath = "/Users/beik/Desktop/" + district_name + " Final Report.xlsx"
newfile.save(outpath)
print("Your file is saved at " + outpath)
print("Done")

os.remove('/Users/beik/Desktop/' + district_name + ' MergedFile.xlsx')
#os.remove('/Users/beik/Desktop/' + district_name + ' Vuln Count.xlsx')
