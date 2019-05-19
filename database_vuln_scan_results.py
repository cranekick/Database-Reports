#!/usr/local/bin/python3

import mysql.connector
import pandas as pd
from sys import exit
from subprocess import Popen
from time import time
# import xlsxwriter
import openpyxl
from openpyxl import load_workbook
# from openpyxl.chart import PieChart3D, Reference, Series
from os import remove

start_time = time()
# This is a dependency, we are creating a background ssh tunnel, this is to be able to access the AlienVault DB
# The REQUIREMENT is the your public key is in the known_hosts on sec-mobile-one, use ssh-copy-id root@sec-mobile-one
print("Starting background process. ")
ssh_magic = Popen(['ssh', '-N', '-L', '3306:127.0.0.1:3306', 'sec-mobile-one'], shell=False)

# This is for the name of the files that are created.
district_name = input("What is the name of the district? ")

# This is the query to get the Vulnerability Results from the Database
query = ("""SELECT vjobs.name as Location, hostname as Hostname, hostip as 'Host IP', service as Service, 
        scriptid as 'Vuln ID', vnp.cve_id as CVEs, vnres.risk as 'Risk Level', vnp.name, vnf.name, vnc.name,
        vnp.copyright, vnp.version, msg 
        FROM alienvault.vuln_nessus_results vnres 
        join vuln_jobs vjobs on vjobs.report_id = vnres.report_id
        join vuln_nessus_plugins vnp on vnres.scriptid = vnp.id
        join vuln_nessus_category vnc on vnc.id = vnp.category
        join vuln_nessus_family vnf on vnf.id = vnp.family
        order by vnres.risk ASC""")


# This is to get the total numbers per site as scans are based on physical locations
query1 = ("""SELECT vjobs.name as Location, vSerious as Critical, vHigh as High, vMed as Medium, vLow as Low
        from vuln_nessus_report_stats vnrstats
        join vuln_jobs vjobs on vjobs.id = vnrstats.name""")


# Accessing the AlienVault MySQL Database
def db_call(query):
    try:
        cnx = mysql.connector.connect(user='svcpython', password='e6$E7sQBUegYayoU',
                                      host='localhost',
                                      database='alienvault')
    except Exception as e:
        print("I am unable to connect to the MySQL Instance, " + str(e))
        exit(1)

    # noinspection PyUnboundLocalVariable
    cursor = cnx.cursor()
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    cnx.close()
    print("Retrieving results from the Database.")
    return results


# GLOBAL
header_row = ['Location', 'Hostname', 'Host IP', 'Service', 'Vuln ID', 'CVE', 'Risk Level', 'Vulnerability',
              '9', '10', '11', '12', 'Blob']
header_row1 = ['Location', 'Critical', 'High', 'Medium', 'Low']
df = pd.DataFrame(db_call(query), columns=header_row)
df = df.astype(str)
df1 = pd.DataFrame(db_call(query1), columns=header_row1)
df2 = pd.DataFrame(df[['Vulnerability','Risk Level']])


def df_rm_garbage():
    # The following function removes 'SCHEDULED - ' and any trailing digits (1 - 10+) in the Locations Column
    # of the DataFrame
    print("Removing 'SCHEDULED - ' and number from the Location")
    df['Location'] = df['Location'].replace(to_replace=r'SCHEDULED - ', value='', regex=True)
    df['Location'] = df['Location'].replace(to_replace=r'(\d+)', value='', regex=True)
    df1['Location'] = df1['Location'].replace(to_replace=r'SCHEDULED - ', value='', regex=True)
    df1['Location'] = df1['Location'].replace(to_replace=r'(\d+)', value='', regex=True)
    return df, df1



def df_vuln_counts():
    # The following replaces the the numerical structure of the 'Risk Level' column with a string based structure.
    # It also groups vulnerabilites by quantity in order to determine what the most prevalent vulns are within the
    # environment. It finally sorts it based on alphabetical order first then by quantity in descending fashion.
    df2['Risk Level'].replace(to_replace=['1', '2', '3', '4', '5', '6', '7', '8'],
                             value=['Critical', 'High', 'Medium', 'Medium/Low', 'Low/Medium', 'Low', 'Info',
                                    'Exceptions'], inplace=True)
    df3 = df2[(df2['Risk Level'] == 'Critical') | (df2['Risk Level'] == 'High') | \
                      (df2['Risk Level'] == 'Medium')].groupby(['Vulnerability', 'Risk Level']) \
                       .size().reset_index(name='Quantity')
    df3.sort_values(by=['Risk Level', 'Quantity'], ascending=[True, False], inplace=True)
    return df3


def df_shenanigans():
    print("Replacing 'Risk Level' integer with string.")
    print("Creating AlienVault Excel view for Columns.")
    df['Vulnerability'] = df['Vulnerability'] + '\n' + 'Family name: ' + df['9'] + '\n' + 'Category: ' + df['10'] + \
                          '\n' + 'Copyright: ' + df['11'] + '\n' + 'Version: ' + df['12']
    df['Risk Level'].replace(to_replace=['1', '2', '3', '4', '5', '6', '7', '8'],
                             value=['Critical', 'High', 'Medium', 'Medium/Low', 'Low/Medium', 'Low', 'Info',
                                    'Exceptions'], inplace=True)
    return df


def df_extract():
    # Extracting values from the AlienVault Database as it is a hodgepodge of information, this presents the information
    # like when you download the Excel
    print("Extracting fields and placing the data in AlienVault Excel view for Columns")
    df['CVSS'] = df['Blob'].str.extract('Score:\s(\d.+)', expand=True)
    df['Observation'] = df['Blob'].str.extract('Summary:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['Remediation'] = df['Blob'].str.extract('Solution:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['insight'] = df['Blob'].str.extract('Insight:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['references'] = df['Blob'].str.extract('(References:\n\n(?:[^\n][\n]?)+)', expand=False)
    df['Consequences'] = df['insight'] + '\n' + df['references']
    df['Test Output'] = df['Blob'].str.extract('Result:\n\n((?:[^\n][\n]?)+)', expand=False)
    df['Operating System/Software'] = df['Blob'].str.extract('OS:\n\n((?:[^\n][\n]?)+)', expand=False)
    return df


def df_drop_columns():
    # Removing Columns that are no longer necessary
    print("Cleaning up the table structure. ")
    df.drop(['9', '10', '11', '12'], inplace=True, axis=1)
    df.drop(['Blob', 'insight', 'references'], inplace=True, axis=1)
    return df


def df_apply_sort():
    # This sort the DataFrame by CVSS Score. 10 - 0 in descending order. This is for precendence sake
    print("Sorting results by CVSS.")
    df[['CVSS']] = df[['CVSS']].apply(pd.to_numeric)
    df.sort_values(by='CVSS', ascending=False, inplace=True)
    return df


def main():
    db_call(query)
    db_call(query1)
    df_rm_garbage()
    df_vuln_counts()
    df_shenanigans()
    df_extract()
    df_drop_columns()
    df_apply_sort()


if __name__ == "__main__":
    main()


df = df[['Location', 'Hostname', 'Host IP', 'Service', 'Risk Level', 'CVSS', 'Vuln ID', 'CVE', 'Vulnerability',
         'Observation', 'Remediation', 'Consequences', 'Test Output', 'Operating System/Software']]

# writer = pd.ExcelWriter('/Users/beik/Desktop/Initial Data.xlsx', engine='xlsxwriter')




critical_df = df.where(df['Risk Level'] == 'Critical')
processed_crits = critical_df.dropna(how='all')
high_df = df.where(df['Risk Level'] == 'High')
processed_highs = high_df.dropna(how='all')
medium_df = df.where(df['Risk Level'] == 'Medium')
processed_mediums = medium_df.dropna(how='all')
low_df = df.where(df['Risk Level'] == 'Low')
processed_lows = low_df.dropna(how='all')
info_df = df.where(df['Risk Level'] == 'Info')
processed_infos = info_df.dropna(how='all')


num_of_crits = processed_crits.shape[0]
num_of_highs = processed_highs.shape[0]
num_of_mediums = processed_mediums.shape[0]
num_of_lows = processed_lows.shape[0]

def ex_data():
    from xlsxwriter import Workbook
    file = Workbook('/Users/beik/Desktop/Initial Report.xlsx')
    ws0 = file.add_worksheet('Executive Summary')
    headings = ['Severity', 'Quantity']
    data = [
        ['Critical', 'High', 'Medium', 'Low'],
        [num_of_crits, num_of_highs, num_of_mediums, num_of_lows]
            ]
    ws0.write_row('A1', headings)
    ws0.write_column('A2', data[0])
    ws0.write_column('B2', data[1])

    chart2 = file.add_chart({'type': 'pie'})

# Configure the series and add user defined segment colors.
    chart2.add_series({
        'name': 'Vulnerabiliities by Severity',
        'categories': "='Executive Summary'!$A$2:$A$5",
        'values':     "='Executive Summary'!$B$2:$B$5",
        'data_labels': {'percentage': True},
        'points': [
            {'fill': {'color': '#FF00FF'}},
            {'fill': {'color': '#FF0000'}},
            {'fill': {'color': '#FFC000'}},
            {'fill': {'color': '#FFFF00'}}
     ],
    })

# Add a title.
    chart2.set_title({'name': 'Vulnerabilities by Severity'})

# Insert the chart into the worksheet (with an offset).
    ws0.insert_chart('D1', chart2, {'x_offset': 25, 'y_offset': 10})

    file.close()

ex_data()

# This needs to happen after ex_data
book = load_workbook('/Users/beik/Desktop/Initial Report.xlsx')
writer = pd.ExcelWriter('/Users/beik/Desktop/Initial Report.xlsx', engine='openpyxl', mode='a')

processed_crits.to_excel(writer, sheet_name='Critical')
processed_highs.to_excel(writer, sheet_name='High')
processed_mediums.to_excel(writer, sheet_name='Medium')
processed_lows.to_excel(writer, sheet_name='Low')
processed_infos.to_excel(writer, sheet_name='Info')
# df.to_excel(writer, sheet_name='Raw Data')
# df1.to_excel(writer, sheet_name='Location Breakdown') Still need to get uniq and add values
df_vuln_counts().to_excel(writer, sheet_name='Unique by Severity')
writer.save()
writer.close()

final_wb = load_workbook('/Users/beik/Desktop/Initial Report.xlsx')

ws0 = final_wb['Executive Summary']
ws1 = final_wb['Critical']
ws2 = final_wb['High']
ws3 = final_wb['Medium']
ws4 = final_wb['Low']
ws5 = final_wb['Info']
ws6 = final_wb['Unique by Severity']

vuln_data = [ws1, ws2, ws3, ws4, ws5]

ws1.sheet_properties.tabColor='FF00FF' # Critical
ws2.sheet_properties.tabColor='FF0000' # High
ws3.sheet_properties.tabColor='FFC000' # Medium
ws4.sheet_properties.tabColor='FFFF00' # Low
ws5.sheet_properties.tabColor='00B050' # Info

for row in ws1.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws3.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws4.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested

for row in ws5.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True) # Tested
        
        
def format_column(ws):
    ws.column_dimensions["B"].width = 14.50
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 5
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 38
    ws.column_dimensions["L"].width = 25
    ws.column_dimensions["M"].width = 25
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["O"].width = 30
    return ws

def handle_col(*sheets):
    return [format_column(sheet) for sheet in sheets]


proper_width = handle_col(*[ws1, ws2, ws3, ws4, ws5])

ws6.column_dimensions["B"].width = 90

a = ws1['A2']
b = ws2['A2']
c = ws3['A2']
d = ws4['A2']
e = ws5['A2']
f = ws6['A2']
ws1.freeze_panes = a
ws2.freeze_panes = b
ws3.freeze_panes = c
ws4.freeze_panes = d
ws5.freeze_panes = e
ws6.freeze_panes = f

def hide_column(ws, column_id):
    if isinstance(column_id, int):
        assert column_id >= 1, "Column numbers must be 1 or greater"
        column_id = openpyxl.utils.get_column_letter(column_id)
    column_dimension = ws.column_dimensions[column_id]
    column_dimension.hidden = True

def spy():
    hide_column(ws1, 1)
    hide_column(ws2, 1)
    hide_column(ws3, 1)
    hide_column(ws4, 1)
    hide_column(ws5, 1)
    hide_column(ws6, 1)

spy()

final_wb.save('/Users/beik/Desktop/' + district_name + ' Final Report.xlsx')
final_wb.close()

rm_initial_report = remove('/Users/beik/Desktop/Initial Report.xlsx')

ssh_magic.kill()
print("The background process is now terminated!")
end_time = time()
running_time = (end_time - start_time)
running_minutes = running_time / 60
print("It took " + str(running_minutes) + " minutes to process.")
